from __future__ import annotations

import copy
import re
from datetime import datetime, time
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter

AC_RE = re.compile(r"^(?:XA-[A-Z0-9]{3}|N[0-9]{3}[A-Z]{2})$", re.I)


def _txt(v): return "" if v is None else str(v).strip()
def _norm(v): return re.sub(r"\s+", " ", _txt(v).replace("_x000D_", " ").replace("\r", " ")).upper().strip()
def _is_ac(v): return bool(AC_RE.fullmatch(_txt(v).upper()))
def _base(v): return re.sub(r"\s+(?:ENG\s*[12]|LH|RH)$", "", _norm(v)).strip()


def _load(data, data_only=False, read_only=False):
    return load_workbook(BytesIO(data), data_only=data_only, read_only=read_only, keep_links=True)


def _find(ws, value, col=1):
    target = _norm(value)
    for r in range(1, ws.max_row + 1):
        if _norm(ws.cell(r, col).value) == target: return r
    raise ValueError(f"No se encontro '{value}' en la hoja {ws.title}.")


def _program(data):
    wb = _load(data, data_only=True, read_only=True); ws = wb[wb.sheetnames[0]]
    blocks=[]; current=None
    for values in ws.iter_rows(values_only=True):
        a,b,c,d=(list(values)+[None]*4)[:4]
        if _is_ac(a):
            current={"ac":_txt(a).upper(),"wo":_txt(b),"tasks":[]}; blocks.append(current)
            if _txt(c): current["tasks"].append((_txt(c),_txt(d) or _txt(c)))
        elif current and _txt(c): current["tasks"].append((_txt(c),_txt(d) or _txt(c)))
    wb.close()
    if not blocks: raise ValueError("Programa Procesado no contiene matriculas validas.")
    return blocks


def _flights(wb):
    if "Sheet1" not in wb.sheetnames: return []
    ws=wb["Sheet1"]
    first=next((r for r in range(1,ws.max_row+1) if _is_ac(ws.cell(r,1).value)),None)
    if first is None: return []
    return [{"ac":_txt(ws.cell(r,1).value).upper(),"fl":ws.cell(r,2).value,"arr":ws.cell(r,3).value,"dept":ws.cell(r,4).value}
            for r in range(first,ws.max_row+1) if _is_ac(ws.cell(r,1).value)]


def _minutes(v):
    if isinstance(v,datetime): v=v.time()
    if isinstance(v,time): return v.hour*60+v.minute
    if isinstance(v,(int,float)): return int(round((float(v)%1)*1440))%1440
    return None


def _copy_style(ws,src_row,dst_row):
    ws.row_dimensions[dst_row].height=ws.row_dimensions[src_row].height
    for c in range(1,ws.max_column+1):
        src,dst=ws.cell(src_row,c),ws.cell(dst_row,c)
        if isinstance(src,MergedCell) or isinstance(dst,MergedCell): continue
        dst._style=copy.copy(src._style); dst.number_format=src.number_format
        dst.alignment=copy.copy(src.alignment); dst.protection=copy.copy(src.protection)


def _hm(wb):
    exact,base={},{}
    if "HM" in wb.sheetnames:
        for card,_desc,mh in wb["HM"].iter_rows(min_row=2,max_col=3,values_only=True):
            if _txt(card): exact.setdefault(_norm(card),mh); base.setdefault(_base(card),mh)
    return exact,base


def _shift_summary(ws,summary,extra):
    if extra<=0: return summary
    merges=[]
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row>=summary:
            merges.append((rng.min_col,rng.min_row,rng.max_col,rng.max_row)); ws.unmerge_cells(str(rng))
    formulas={}
    for r in range(summary,ws.max_row+1):
        for c in range(1,ws.max_column+1):
            v=ws.cell(r,c).value
            if isinstance(v,str) and v.startswith("="): formulas[(r,c)]=v
    ws.insert_rows(summary,extra)
    for a,b,c,d in merges: ws.merge_cells(start_row=b+extra,start_column=a,end_row=d+extra,end_column=c)
    threshold=summary-2
    for (old_r,c),formula in formulas.items():
        new_r=old_r+extra; old=f"{get_column_letter(c)}{old_r}"; new=f"{get_column_letter(c)}{new_r}"
        try: translated=Translator(formula,origin=old).translate_formula(new)
        except Exception: translated=formula
        def adjust(m):
            n=int(m.group(2)); return m.group(1)+str(n+extra if n>=threshold else n)
        ws.cell(new_r,c).value=re.sub(r"(\$?[A-Z]{1,3}\$)(\d+)",adjust,translated)
    return summary+extra


def generar(template_bytes, program_bytes):
    wb=_load(template_bytes)
    if "GDL" not in wb.sheetnames: raise ValueError("Daily Check Plantilla debe contener la hoja GDL.")
    ws=wb["GDL"]; blocks=_program(program_bytes); flights=_flights(wb); exact,base=_hm(wb)
    start=_find(ws,"A/C")+1; summary=_find(ws,"RON AC")
    by_ac={b["ac"]:b for b in blocks}; latest={}
    for f in flights: latest[f["ac"]]=f
    missing=[b for b in blocks if b["ac"] not in latest]
    records=[]
    for b in missing: records.append((b,{"ac":b["ac"],"fl":"STORAGE","arr":None,"dept":None}))
    used=set()
    for f in sorted(flights,key=lambda x: 99999 if _minutes(x["arr"]) is None else (_minutes(x["arr"])-1080)%1440):
        b=by_ac.get(f["ac"])
        if b and f is latest[f["ac"]] and f["ac"] not in used: records.append((b,f)); used.add(f["ac"])
        else: records.append((None,f))
    needed=sum(1 if b is None else max(1,len(b["tasks"])) for b,_ in records)
    summary=_shift_summary(ws,summary,max(0,needed-(summary-start)))
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row>=start and rng.max_row<summary: ws.unmerge_cells(str(rng))
    for r in range(start,summary):
        for c in range(1,17):
            if not isinstance(ws.cell(r,c),MergedCell): ws.cell(r,c).value=None
    row=start
    for block,flight in records:
        if block is None:
            _copy_style(ws,start,row); ws.cell(row,1).value=flight["ac"]; ws.cell(row,2).value=flight["fl"]
            ws.cell(row,3).value=flight["arr"]; ws.cell(row,4).value=flight["dept"]
            a,d=_minutes(flight["arr"]),_minutes(flight["dept"])
            ws.cell(row,5).value="TRANSIT CHECK / RON" if a is not None and d is not None and d<a else "TRANSIT CHECK"; row+=1; continue
        tasks=block["tasks"] or [("","")]; first=row; last=row+len(tasks)-1
        for card,desc in tasks:
            _copy_style(ws,start,row); ws.cell(row,6).value=card or None; ws.cell(row,7).value=desc or None
            mh=exact.get(_norm(card)) if card else None
            if mh is None and card: mh=base.get(_base(card))
            ws.cell(row,10).value=mh; row+=1
        ws.cell(first,1).value=block["ac"]; ws.cell(first,2).value=flight["fl"]
        ws.cell(first,3).value=flight["arr"]; ws.cell(first,4).value=flight["dept"]; ws.cell(first,5).value=block["wo"]
        if last>first:
            for c in range(1,6): ws.merge_cells(start_row=first,start_column=c,end_row=last,end_column=c)
    wb.calculation.fullCalcOnLoad=True; wb.calculation.forceFullCalc=True
    out=BytesIO(); wb.save(out); wb.close()
    return out.getvalue(),{"matriculas":len(blocks),"tareas":sum(len(b["tasks"]) for b in blocks)}

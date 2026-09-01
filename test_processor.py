from pathlib import Path
from openpyxl import load_workbook
from processor import generate_combined

BASE = Path(__file__).resolve().parent
programa = BASE / "PROGRAMA_PROCESADO.xlsx"
plantilla = BASE / "DAILY CHECK PLANTILLA.xlsx"
esperado = BASE / "DAILY CHECK COMBINADO.xlsx"
salida = BASE / "RESULTADO_PRUEBA.xlsx"

for archivo in (programa, plantilla):
    if not archivo.exists():
        raise FileNotFoundError(f"Falta {archivo.name}")

generate_combined(programa, plantilla, salida)
wb_out = load_workbook(salida, data_only=False)
wb_tpl = load_workbook(plantilla, data_only=False)
assert wb_out.sheetnames == wb_tpl.sheetnames
assert "GDL" in wb_out.sheetnames
assert wb_out["GDL"].max_row > wb_tpl["GDL"].max_row


program_ac = {str(wb_out["GDL"].cell(r, 1).value).strip() for r in range(1, wb_out["GDL"].max_row + 1) if wb_out["GDL"].cell(r, 1).value}
assert "XA-VRQ" in program_ac

print(f"Prueba terminada: {salida.name}")

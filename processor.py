from __future__ import annotations

import copy
import io
import re
from dataclasses import dataclass
from datetime import datetime, time, timedelta
from pathlib import Path
from typing import BinaryIO, Iterable

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter


@dataclass
class Task:
    task: str
    description: str


@dataclass
class ProgramGroup:
    ac: str
    wo: str
    tasks: list[Task]


@dataclass
class Flight:
    row: int
    ac: str
    fl: object
    arr: object
    dept: object


def text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def norm_ac(value) -> str:
    return re.sub(r"\s+", "", text(value).upper())


def as_seconds(value) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        value = value.time()
    if isinstance(value, time):
        return value.hour * 3600 + value.minute * 60 + value.second
    if isinstance(value, (int, float)):
        fraction = float(value) % 1
        return int(round(fraction * 86400)) % 86400
    raw = text(value)
    for fmt in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M:%S %p"):
        try:
            parsed = datetime.strptime(raw, fmt)
            return parsed.hour * 3600 + parsed.minute * 60 + parsed.second
        except ValueError:
            pass
    return None


def elapsed_seconds(arr, dept) -> int:
    a, d = as_seconds(arr), as_seconds(dept)
    if a is None or d is None:
        return -1
    return d - a if d >= a else 86400 - a + d


def is_ron(arr, dept) -> bool:
    a, d = as_seconds(arr), as_seconds(dept)
    return a is not None and d is not None and d <= a


def parse_program(source) -> list[ProgramGroup]:
    wb = load_workbook(source, data_only=False)
    ws = wb["Programa"] if "Programa" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = {text(ws.cell(1, c).value).upper(): c for c in range(1, ws.max_column + 1)}
    required = {"AC", "WO", "TASK", "DESCRIPTION"}
    if not required.issubset(headers):
        raise ValueError("PROGRAMA PROCESADO debe contener AC, WO, TASK y DESCRIPTION en la fila 1.")

    groups: list[ProgramGroup] = []
    current: ProgramGroup | None = None
    current_ac = current_wo = ""
    for row in range(2, ws.max_row + 1):
        ac_cell = norm_ac(ws.cell(row, headers["AC"]).value)
        wo_cell = text(ws.cell(row, headers["WO"]).value)
        task = text(ws.cell(row, headers["TASK"]).value)
        description = text(ws.cell(row, headers["DESCRIPTION"]).value)
        if ac_cell:
            current_ac = ac_cell
        if wo_cell:
            current_wo = wo_cell
        if not task and not description:
            continue
        if current is None or current.ac != current_ac or current.wo != current_wo:
            current = ProgramGroup(current_ac, current_wo, [])
            groups.append(current)
        current.tasks.append(Task(task, description))
    return [g for g in groups if g.ac and g.tasks]


def find_schedule(ws) -> tuple[int, int]:
    header_row = None
    for r in range(1, min(ws.max_row, 20) + 1):
        values = [text(ws.cell(r, c).value).upper() for c in range(1, 9)]
        if values[:4] == ["A/C", "FL", "ARR", "DEPT"]:
            header_row = r
            break
    if header_row is None:
        raise ValueError("No se encontró el encabezado A/C, FL, ARR, DEPT en la hoja GDL.")
    first_data = header_row + 1
    start = first_data
    # RTO y STORAGE forman un bloque fijo al inicio. No deben convertirse en tránsitos.
    while start <= ws.max_row:
        operation = text(ws.cell(start, 2).value).upper()
        if operation not in {"RTO", "STORAGE", ""}:
            break
        start += 1
    end = None
    for r in range(start, ws.max_row + 1):
        if text(ws.cell(r, 1).value).upper() == "RON AC":
            end = r - 1
            break
    if end is None:
        raise ValueError("No se encontró la fila RON AC que cierra el bloque de vuelos.")
    return start, end


def copy_cell(src, dst, translate_from: str | None = None):
    value = src.value
    if translate_from and isinstance(value, str) and value.startswith("="):
        try:
            value = Translator(value, origin=translate_from).translate_formula(dst.coordinate)
        except Exception:
            pass
    dst.value = value
    if src.has_style:
        dst._style = copy.copy(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    dst.font = copy.copy(src.font)
    dst.fill = copy.copy(src.fill)
    dst.border = copy.copy(src.border)
    dst.alignment = copy.copy(src.alignment)
    dst.protection = copy.copy(src.protection)
    if src.hyperlink:
        dst._hyperlink = copy.copy(src.hyperlink)
    if src.comment:
        dst.comment = copy.copy(src.comment)


def copy_row_style(ws, source_row: int, target_row: int, max_col: int):
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    ws.row_dimensions[target_row].hidden = ws.row_dimensions[source_row].hidden
    for c in range(1, max_col + 1):
        src, dst = ws.cell(source_row, c), ws.cell(target_row, c)
        if src.has_style:
            dst._style = copy.copy(src._style)
        dst.number_format = src.number_format
        dst.alignment = copy.copy(src.alignment)
        dst.protection = copy.copy(src.protection)


def unmerge_intersecting(ws, start_row: int, end_row: int):
    for rg in list(ws.merged_cells.ranges):
        if not (rg.max_row < start_row or rg.min_row > end_row):
            ws.unmerge_cells(str(rg))



def append_storage_tasks(ws, groups: list[ProgramGroup]) -> set[int]:
    """Agrega tareas del programa a bloques STORAGE existentes y devuelve grupos consumidos."""
    consumed: set[int] = set()
    by_ac: dict[str, list[tuple[int, ProgramGroup]]] = {}
    for index, group in enumerate(groups):
        by_ac.setdefault(group.ac, []).append((index, group))

    row = 3
    while row <= ws.max_row:
        ac = norm_ac(ws.cell(row, 1).value)
        if text(ws.cell(row, 2).value).upper() != "STORAGE" or ac not in by_ac:
            row += 1
            continue
        block_start = row
        block_end = row
        while block_end + 1 <= ws.max_row and not norm_ac(ws.cell(block_end + 1, 1).value):
            block_end += 1
        additions = [(idx, task) for idx, group in by_ac[ac] for task in group.tasks]
        if not additions:
            row = block_end + 1
            continue

        insert_at = block_end + 1
        count = len(additions)
        merges = [copy.copy(rg) for rg in ws.merged_cells.ranges]
        for rg in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(rg))
        ws.insert_rows(insert_at, count)
        for rg in merges:
            if rg.min_row >= insert_at:
                ws.merge_cells(start_row=rg.min_row + count, start_column=rg.min_col,
                               end_row=rg.max_row + count, end_column=rg.max_col)
            elif rg.max_row >= block_start and rg.min_row <= block_end:
                # El bloque STORAGE crece para incluir las tareas adicionales.
                ws.merge_cells(start_row=rg.min_row, start_column=rg.min_col,
                               end_row=rg.max_row + count, end_column=rg.max_col)
            else:
                ws.merge_cells(str(rg))

        for offset, (idx, task) in enumerate(additions):
            target = insert_at + offset
            copy_row_style(ws, block_end, target, ws.max_column)
            ws.cell(target, 6).value = task.task
            ws.cell(target, 7).value = task.description
            consumed.add(idx)
        row = insert_at + count
    return consumed

def generate_combined(program_source, template_source, output_path: str | Path | None = None) -> bytes:
    groups = parse_program(program_source)
    wb = load_workbook(template_source, data_only=False)
    if "GDL" not in wb.sheetnames:
        raise ValueError("DAILY CHECK PLANTILLA debe contener la hoja GDL.")
    ws = wb["GDL"]
    consumed_storage = append_storage_tasks(ws, groups)
    groups = [group for index, group in enumerate(groups) if index not in consumed_storage]
    start, end = find_schedule(ws)

    flights = [
        Flight(r, norm_ac(ws.cell(r, 1).value), ws.cell(r, 2).value, ws.cell(r, 3).value, ws.cell(r, 4).value)
        for r in range(start, end + 1)
        if norm_ac(ws.cell(r, 1).value)
    ]
    by_ac: dict[str, list[int]] = {}
    for i, flight in enumerate(flights):
        by_ac.setdefault(flight.ac, []).append(i)

    assigned: dict[int, list[ProgramGroup]] = {}
    for group in groups:
        candidates = by_ac.get(group.ac, [])
        if not candidates:
            continue
        best = max(candidates, key=lambda i: elapsed_seconds(flights[i].arr, flights[i].dept))
        assigned.setdefault(best, []).append(group)

    rendered: list[tuple[Flight, list[ProgramGroup] | None]] = []
    for i, flight in enumerate(flights):
        rendered.append((flight, assigned.get(i)))

    old_summary_start = end + 1
    old_max_row = ws.max_row
    old_merges = [copy.copy(rg) for rg in ws.merged_cells.ranges if rg.min_row >= old_summary_start]
    summary_snapshot = []
    for r in range(old_summary_start, old_max_row + 1):
        row = []
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            row.append((cell.value, copy.copy(cell._style), copy.copy(cell.font), copy.copy(cell.fill),
                        copy.copy(cell.border), copy.copy(cell.alignment), copy.copy(cell.protection),
                        cell.number_format, copy.copy(cell.comment), copy.copy(cell.hyperlink)))
        summary_snapshot.append((ws.row_dimensions[r].height, ws.row_dimensions[r].hidden, row))

    total_rows = sum(max(1, sum(len(g.tasks) for g in gs)) if gs else 1 for _, gs in rendered)
    delta = total_rows - len(flights)
    unmerge_intersecting(ws, start, old_max_row)
    if delta > 0:
        ws.insert_rows(old_summary_start, delta)

    for r in range(start, start + total_rows):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None
        copy_row_style(ws, start, r, ws.max_column)

    out_row = start
    programmed_count = transit_count = ron_count = 0
    for flight, flight_groups in rendered:
        if not flight_groups:
            copy_row_style(ws, flight.row, out_row, ws.max_column)
            ws.cell(out_row, 1).value = flight.ac
            ws.cell(out_row, 2).value = flight.fl
            ws.cell(out_row, 3).value = flight.arr
            ws.cell(out_row, 4).value = flight.dept
            label = "TRANSIT CHECK / RON" if is_ron(flight.arr, flight.dept) else "TRANSIT CHECK"
            ws.cell(out_row, 5).value = label
            ws.merge_cells(start_row=out_row, start_column=5, end_row=out_row, end_column=8)
            transit_count += 1
            ron_count += int(label.endswith("RON"))
            out_row += 1
            continue

        programmed_count += 1
        task_rows = [(g, task) for g in flight_groups for task in g.tasks]
        block_start = out_row
        block_end = out_row + len(task_rows) - 1
        copy_row_style(ws, flight.row, block_start, ws.max_column)
        for r in range(block_start, block_end + 1):
            copy_row_style(ws, flight.row, r, ws.max_column)
        ws.cell(block_start, 1).value = flight.ac
        ws.cell(block_start, 2).value = flight.fl
        ws.cell(block_start, 3).value = flight.arr
        ws.cell(block_start, 4).value = flight.dept
        wo_values = []
        for g in flight_groups:
            if g.wo and g.wo not in wo_values:
                wo_values.append(g.wo)
        ws.cell(block_start, 5).value = "\n".join(wo_values)
        for col in range(1, 6):
            if block_end > block_start:
                ws.merge_cells(start_row=block_start, start_column=col, end_row=block_end, end_column=col)
        for r, (_, task) in enumerate(task_rows, start=block_start):
            ws.cell(r, 6).value = task.task
            ws.cell(r, 7).value = task.description
        out_row = block_end + 1

    new_summary_start = start + total_rows
    for offset, (height, hidden, rowdata) in enumerate(summary_snapshot):
        target_r = new_summary_start + offset
        ws.row_dimensions[target_r].height = height
        ws.row_dimensions[target_r].hidden = hidden
        for c, data in enumerate(rowdata, start=1):
            value, style, font, fill, border, alignment, protection, numfmt, comment, hyperlink = data
            dst = ws.cell(target_r, c)
            if isinstance(value, str) and value.startswith("="):
                origin = f"{get_column_letter(c)}{old_summary_start + offset}"
                try:
                    value = Translator(value, origin=origin).translate_formula(dst.coordinate)
                except Exception:
                    pass
            dst.value = value
            dst._style, dst.font, dst.fill, dst.border = style, font, fill, border
            dst.alignment, dst.protection, dst.number_format = alignment, protection, numfmt
            dst.comment, dst._hyperlink = comment, hyperlink

    for rg in old_merges:
        shift = new_summary_start - old_summary_start
        ws.merge_cells(start_row=rg.min_row + shift, start_column=rg.min_col,
                       end_row=rg.max_row + shift, end_column=rg.max_col)

    ron_row = new_summary_start
    ws.cell(ron_row, 5).value = programmed_count
    # Update visible summary counters when their labels are present.
    for r in range(new_summary_start, ws.max_row + 1):
        label = text(ws.cell(r, 1).value).upper()
        if label == "PROGRAMADOS": ws.cell(r, 2).value = programmed_count
        elif label == "TRANSITOS": ws.cell(r, 2).value = transit_count
    ws.print_area = f"A1:M{new_summary_start - 3}" if new_summary_start > 3 else ws.print_area
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    output = io.BytesIO()
    wb.save(output)
    data = output.getvalue()
    if output_path:
        Path(output_path).write_bytes(data)
    return data

from __future__ import annotations

import copy
import re
from datetime import date, datetime, time
from io import BytesIO

import streamlit as st
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter

AC_RE = re.compile(r"^(?:XA-[A-Z0-9]{3}|N[0-9]{3}[A-Z]{2})$", re.I)


def txt(value):
    return "" if value is None else str(value).strip()


def norm(value):
    return re.sub(r"\s+", " ", txt(value).replace("_x000D_", " ").replace("\r", " ")).upper().strip()


def es_ac(value):
    return bool(AC_RE.fullmatch(txt(value).upper()))


def tarea_base(value):
    return re.sub(r"\s+(?:ENG\s*[12]|LH|RH)$", "", norm(value)).strip()


def cargar_excel(uploaded, data_only=False, read_only=False):
    uploaded.seek(0)
    return load_workbook(uploaded, data_only=data_only, read_only=read_only, keep_links=True)


def leer_programa(uploaded):
    wb = cargar_excel(uploaded, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    bloques, actual = [], None
    for values in ws.iter_rows(values_only=True):
        a, b, c, d = (list(values) + [None] * 4)[:4]
        if es_ac(a):
            actual = {"ac": txt(a).upper(), "wo": txt(b), "tareas": []}
            bloques.append(actual)
            if txt(c):
                actual["tareas"].append((txt(c), txt(d) or txt(c)))
        elif actual and txt(c):
            actual["tareas"].append((txt(c), txt(d) or txt(c)))
    wb.close()
    if not bloques:
        raise ValueError("El programa procesado no contiene matriculas reconocibles en la columna AC.")
    return bloques


def buscar_fila(ws, text_value, column=1):
    target = norm(text_value)
    for row in range(1, ws.max_row + 1):
        if norm(ws.cell(row, column).value) == target:
            return row
    raise ValueError(f"No se encontro '{text_value}' en la hoja {ws.title}.")


def leer_vuelos(wb):
    if "Sheet1" not in wb.sheetnames:
        return []
    ws = wb["Sheet1"]
    first = next((r for r in range(1, ws.max_row + 1) if es_ac(ws.cell(r, 1).value)), None)
    if first is None:
        return []
    result = []
    for row in range(first, ws.max_row + 1):
        ac = txt(ws.cell(row, 1).value).upper()
        if es_ac(ac):
            result.append({"ac": ac, "fl": ws.cell(row, 2).value, "arr": ws.cell(row, 3).value, "dept": ws.cell(row, 4).value})
    return result


def minutes(value):
    if isinstance(value, datetime):
        value = value.time()
    if isinstance(value, time):
        return value.hour * 60 + value.minute
    if isinstance(value, (int, float)):
        return int(round((float(value) % 1) * 1440)) % 1440
    return None


def flight_key(flight):
    m = minutes(flight["arr"])
    return 99999 if m is None else (m - 18 * 60) % 1440


def copy_row_style(ws, source, target):
    ws.row_dimensions[target].height = ws.row_dimensions[source].height
    for col in range(1, ws.max_column + 1):
        src, dst = ws.cell(source, col), ws.cell(target, col)
        if isinstance(src, MergedCell) or isinstance(dst, MergedCell):
            continue
        dst._style = copy.copy(src._style)
        dst.number_format = src.number_format
        dst.alignment = copy.copy(src.alignment)
        dst.protection = copy.copy(src.protection)


def build_hm(wb):
    exact, base = {}, {}
    if "HM" in wb.sheetnames:
        for card, _desc, mh in wb["HM"].iter_rows(min_row=2, max_col=3, values_only=True):
            if txt(card):
                exact.setdefault(norm(card), mh)
                base.setdefault(tarea_base(card), mh)
    return exact, base


def shift_summary(ws, summary_row, extra):
    if extra <= 0:
        return summary_row
    merges = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= summary_row:
            merges.append((rng.min_col, rng.min_row, rng.max_col, rng.max_row))
            ws.unmerge_cells(str(rng))
    formulas = {}
    for row in range(summary_row, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row, col).value
            if isinstance(value, str) and value.startswith("="):
                formulas[(row, col)] = value
    ws.insert_rows(summary_row, extra)
    for min_col, min_row, max_col, max_row in merges:
        ws.merge_cells(start_row=min_row + extra, start_column=min_col, end_row=max_row + extra, end_column=max_col)
    threshold = summary_row - 2
    for (old_row, col), formula in formulas.items():
        new_row = old_row + extra
        old_coord = f"{get_column_letter(col)}{old_row}"
        new_coord = f"{get_column_letter(col)}{new_row}"
        try:
            translated = Translator(formula, origin=old_coord).translate_formula(new_coord)
        except Exception:
            translated = formula
        def adjust(match):
            row_number = int(match.group(2))
            return match.group(1) + str(row_number + extra if row_number >= threshold else row_number)
        translated = re.sub(r"(\$?[A-Z]{1,3}\$)(\d+)", adjust, translated)
        ws.cell(new_row, col).value = translated
    return summary_row + extra


def generate_daily_check(template_file, program_file):
    wb = cargar_excel(template_file)
    if "GDL" not in wb.sheetnames:
        raise ValueError("La plantilla debe contener la hoja GDL.")
    ws = wb["GDL"]
    blocks = leer_programa(program_file)
    program_by_ac = {b["ac"]: b for b in blocks}
    flights = sorted(leer_vuelos(wb), key=flight_key)
    exact_hm, base_hm = build_hm(wb)

    start = buscar_fila(ws, "A/C") + 1
    summary = buscar_fila(ws, "RON AC")

    # Un bloque programado ocupa una fila por tarea. Un transito ocupa una fila.
    latest_flight = {}
    for flight in flights:
        latest_flight[flight["ac"]] = flight
    programmed_in_flights = {ac for ac in program_by_ac if ac in latest_flight}
    missing_programmed = [b for b in blocks if b["ac"] not in programmed_in_flights]
    rows_needed = len([f for f in flights if f["ac"] not in program_by_ac or latest_flight[f["ac"]] is f])
    rows_needed += sum(max(1, len(b["tareas"])) - 1 for b in blocks if b["ac"] in latest_flight)
    rows_needed += sum(max(1, len(b["tareas"])) for b in missing_programmed)

    available = summary - start
    summary = shift_summary(ws, summary, max(0, rows_needed - available))

    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= start and rng.max_row < summary:
            ws.unmerge_cells(str(rng))
    for row in range(start, summary):
        for col in range(1, 17):
            cell = ws.cell(row, col)
            if not isinstance(cell, MergedCell):
                cell.value = None

    records = []
    for block in missing_programmed:
        records.append((block, {"ac": block["ac"], "fl": "STORAGE", "arr": None, "dept": None}))
    used_programmed = set()
    for flight in flights:
        block = program_by_ac.get(flight["ac"])
        if block and flight is latest_flight[flight["ac"]] and flight["ac"] not in used_programmed:
            records.append((block, flight)); used_programmed.add(flight["ac"])
        elif not block or flight is not latest_flight.get(flight["ac"]):
            records.append((None, flight))

    row = start
    no_hm = []
    for block, flight in records:
        if block is None:
            copy_row_style(ws, start, row)
            ws.cell(row, 1).value, ws.cell(row, 2).value = flight["ac"], flight["fl"]
            ws.cell(row, 3).value, ws.cell(row, 4).value = flight["arr"], flight["dept"]
            a, d = minutes(flight["arr"]), minutes(flight["dept"])
            ws.cell(row, 5).value = "TRANSIT CHECK / RON" if a is not None and d is not None and d < a else "TRANSIT CHECK"
            row += 1
            continue
        tasks = block["tareas"] or [("", "")]
        first, last = row, row + len(tasks) - 1
        for card, description in tasks:
            copy_row_style(ws, start, row)
            ws.cell(row, 6).value = card or None
            ws.cell(row, 7).value = description or None
            mh = exact_hm.get(norm(card)) if card else None
            if mh is None and card:
                mh = base_hm.get(tarea_base(card))
            ws.cell(row, 10).value = mh
            if card and mh is None:
                no_hm.append(card)
            row += 1
        ws.cell(first, 1).value, ws.cell(first, 2).value = block["ac"], flight["fl"]
        ws.cell(first, 3).value, ws.cell(first, 4).value = flight["arr"], flight["dept"]
        ws.cell(first, 5).value = block["wo"]
        if last > first:
            for col in range(1, 6):
                ws.merge_cells(start_row=first, start_column=col, end_row=last, end_column=col)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    output = BytesIO(); wb.save(output); wb.close()
    return output.getvalue(), {"matriculas": len(blocks), "tareas": sum(len(b["tareas"]) for b in blocks), "sin_hm": sorted(set(no_hm))}


st.set_page_config(page_title="Daily Check GDL", layout="centered")
st.title("Generador Daily Check GDL")
st.write("Carga los dos archivos de entrada para generar el Daily Check final ordenado.")

template = st.file_uploader("Daily Check Plantilla", type="xlsx")
program = st.file_uploader("Programa Procesado", type="xlsx")
output_date = st.date_input("Fecha para el nombre del archivo", value=date.today())

if st.button("Generar archivo", type="primary", use_container_width=True):
    if template is None or program is None:
        st.warning("Carga Daily Check Plantilla y Programa Procesado.")
    else:
        try:
            with st.spinner("Generando..."):
                data, stats = generate_daily_check(template, program)
            st.session_state["daily_data"] = data
            st.session_state["daily_stats"] = stats
            st.session_state["daily_name"] = f"Daily check GDL {output_date.strftime('%b %d').upper()}.xlsx"
            st.success(f"Listo: {stats['matriculas']} matriculas y {stats['tareas']} tareas.")
        except Exception as exc:
            st.session_state.pop("daily_data", None)
            st.error(f"No se pudo generar el archivo: {exc}")

if "daily_data" in st.session_state:
    st.download_button(
        "Descargar Daily Check final",
        st.session_state["daily_data"],
        file_name=st.session_state["daily_name"],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

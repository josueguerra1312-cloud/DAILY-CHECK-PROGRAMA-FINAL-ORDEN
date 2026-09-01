"""Motor para incorporar un programa procesado en un Daily Check existente."""
from __future__ import annotations

import copy
import re
from io import BytesIO
from pathlib import Path
from typing import BinaryIO, Union

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter

ExcelSource = Union[str, Path, BinaryIO, BytesIO]
AC_RE = re.compile(r"^(?:XA-[A-Z0-9]{3}|N[0-9]{3}[A-Z]{2})$", re.I)


def _txt(value) -> str:
    return "" if value is None else str(value).strip()


def _norm(value) -> str:
    value = _txt(value).replace("_x000D_", " ").replace("\r", " ")
    return re.sub(r"\s+", " ", value).upper().strip()


def _es_matricula(value) -> bool:
    return bool(AC_RE.fullmatch(_txt(value).upper()))


def _tarea_base(value) -> str:
    return re.sub(r"\s+(?:ENG\s*[12]|LH|RH)$", "", _norm(value)).strip()


def _cargar(source: ExcelSource, *, data_only=False, read_only=False):
    if hasattr(source, "seek"):
        source.seek(0)
    return load_workbook(source, data_only=data_only, read_only=read_only, keep_links=True)


def leer_programa(source: ExcelSource) -> list[dict]:
    wb = _cargar(source, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    resultado: list[dict] = []
    actual = None
    for values in ws.iter_rows(values_only=True):
        a, b, c, d = (list(values) + [None] * 4)[:4]
        if _es_matricula(a):
            actual = {"ac": _txt(a).upper(), "wo": _txt(b), "tareas": []}
            resultado.append(actual)
            if _txt(c):
                actual["tareas"].append((_txt(c), _txt(d) or _txt(c)))
        elif actual and _txt(c):
            actual["tareas"].append((_txt(c), _txt(d) or _txt(c)))
    wb.close()
    if not resultado:
        raise ValueError("No se encontraron matriculas validas en el programa procesado.")
    return resultado


def _buscar_fila(ws, texto: str, columna: int = 1) -> int:
    objetivo = _norm(texto)
    for row in range(1, ws.max_row + 1):
        if _norm(ws.cell(row, columna).value) == objetivo:
            return row
    raise ValueError(f"No se encontro '{texto}' en la hoja '{ws.title}'.")


def _copiar_estilo_fila(ws, origen: int, destino: int) -> None:
    ws.row_dimensions[destino].height = ws.row_dimensions[origen].height
    ws.row_dimensions[destino].hidden = ws.row_dimensions[origen].hidden
    for column in range(1, ws.max_column + 1):
        src = ws.cell(origen, column)
        dst = ws.cell(destino, column)
        if isinstance(src, MergedCell) or isinstance(dst, MergedCell):
            continue
        dst._style = copy.copy(src._style)
        dst.number_format = src.number_format
        dst.alignment = copy.copy(src.alignment)
        dst.protection = copy.copy(src.protection)


def _catalogo_hm(wb):
    exacto, base = {}, {}
    if "HM" not in wb.sheetnames:
        return exacto, base
    for card, _desc, mh in wb["HM"].iter_rows(min_row=2, max_col=3, values_only=True):
        if _txt(card):
            exacto.setdefault(_norm(card), mh)
            base.setdefault(_tarea_base(card), mh)
    return exacto, base


def _vuelos(wb):
    if "Sheet1" not in wb.sheetnames:
        return []
    ws = wb["Sheet1"]
    header = None
    for row in range(1, min(ws.max_row, 30) + 1):
        values = [_norm(ws.cell(row, col).value) for col in range(1, 5)]
        if values[:2] in (["A/C", "FL"], ["AC", "FL"]):
            header = row
            break
    if header is None:
        # Algunas plantillas no tienen encabezado visible; acepta la primera fila de matricula.
        first_data = next((row for row in range(1, ws.max_row + 1) if _es_matricula(ws.cell(row, 1).value)), None)
        if first_data is None:
            return []
        header = first_data - 1
    result = []
    for row in range(header + 1, ws.max_row + 1):
        ac = _txt(ws.cell(row, 1).value).upper()
        if _es_matricula(ac):
            result.append((ac, ws.cell(row, 2).value, ws.cell(row, 3).value, ws.cell(row, 4).value))
    return result


def generar_daily_check(plantilla: ExcelSource, programa: ExcelSource) -> tuple[bytes, dict]:
    """Devuelve el Excel final en bytes y un resumen del procesamiento."""
    wb = _cargar(plantilla)
    requeridas = {"GDL", "HM", "Sheet1"}
    faltantes = requeridas.difference(wb.sheetnames)
    if faltantes:
        wb.close()
        raise ValueError("La plantilla no contiene estas hojas: " + ", ".join(sorted(faltantes)))

    ws = wb["GDL"]
    bloques = leer_programa(programa)
    vuelos = _vuelos(wb)
    hm_exacto, hm_base = _catalogo_hm(wb)
    inicio = _buscar_fila(ws, "A/C") + 1
    resumen = _buscar_fila(ws, "RON AC")
    disponibles = resumen - inicio
    necesarias = sum(max(1, len(b["tareas"])) for b in bloques)
    # Retira solo combinaciones del cuerpo operativo; se reconstruyen por matricula.
    for rango in list(ws.merged_cells.ranges):
        if rango.min_row >= inicio and rango.max_row < resumen:
            ws.unmerge_cells(str(rango))
    if necesarias > disponibles:
        extra = necesarias - disponibles
        # Conserva y desplaza correctamente combinaciones y formulas del bloque de resumen.
        merges_inferiores = []
        for rango in list(ws.merged_cells.ranges):
            if rango.min_row >= resumen:
                merges_inferiores.append((rango.min_col, rango.min_row, rango.max_col, rango.max_row))
                ws.unmerge_cells(str(rango))
        formulas = {}
        for row in range(resumen, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                value = ws.cell(row, col).value
                if isinstance(value, str) and value.startswith("="):
                    formulas[(row, col)] = value
        ws.insert_rows(resumen, extra)
        for min_col, min_row, max_col, max_row in merges_inferiores:
            ws.merge_cells(
                start_row=min_row + extra, start_column=min_col,
                end_row=max_row + extra, end_column=max_col
            )
        for (old_row, col), formula in formulas.items():
            new_row = old_row + extra
            old_coord = f"{get_column_letter(col)}{old_row}"
            new_coord = f"{get_column_letter(col)}{new_row}"
            try:
                translated = Translator(formula, origin=old_coord).translate_formula(new_coord)
            except Exception:
                translated = formula
            # openpyxl no desplaza referencias absolutas. Ajusta las que pertenecen
            # al final del cuerpo o al bloque de resumen movido.
            threshold = resumen - 2
            def shift_absolute(match):
                prefix, row_text = match.group(1), match.group(2)
                row_number = int(row_text)
                if row_number >= threshold:
                    row_number += extra
                return prefix + str(row_number)
            translated = re.sub(r"(\$?[A-Z]{1,3}\$)(\d+)", shift_absolute, translated)
            ws.cell(new_row, col).value = translated
        resumen += extra
        disponibles += extra

    # No se elimina ninguna fila. El bloque de resumen permanece debajo del area operativa.
    for row in range(inicio, resumen):
        for column in range(1, 17):
            cell = ws.cell(row, column)
            if not isinstance(cell, MergedCell):
                cell.value = None

    vuelo_por_ac = {}
    for vuelo in vuelos:
        vuelo_por_ac[vuelo[0]] = vuelo  # ultima aparicion, como en el archivo de referencia

    fila = inicio
    tareas_sin_hm = []
    for bloque in bloques:
        tareas = bloque["tareas"] or [("", "")]
        primera = fila
        ultima = fila + len(tareas) - 1
        vuelo = vuelo_por_ac.get(bloque["ac"])

        for card, descripcion in tareas:
            _copiar_estilo_fila(ws, inicio, fila)
            ws.cell(fila, 6).value = card or None
            ws.cell(fila, 7).value = descripcion or None
            mh = hm_exacto.get(_norm(card)) if card else None
            if mh is None and card:
                mh = hm_base.get(_tarea_base(card))
            ws.cell(fila, 10).value = mh
            if card and mh is None:
                tareas_sin_hm.append(card)
            fila += 1

        # Matricula, vuelo, horarios y WO abarcan todas las tareas del avion.
        ws.cell(primera, 1).value = bloque["ac"]
        if vuelo:
            ws.cell(primera, 2).value, ws.cell(primera, 3).value, ws.cell(primera, 4).value = vuelo[1:]
        else:
            ws.cell(primera, 2).value = "STORAGE"
        ws.cell(primera, 5).value = bloque["wo"]
        if ultima > primera:
            for col in range(1, 6):
                ws.merge_cells(start_row=primera, start_column=col, end_row=ultima, end_column=col)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    buffer = BytesIO()
    wb.save(buffer)
    wb.close()
    return buffer.getvalue(), {
        "matriculas": len(bloques),
        "tareas": sum(len(b["tareas"]) for b in bloques),
        "filas": necesarias,
        "sin_hm": sorted(set(tareas_sin_hm)),
    }


def generar_archivo(plantilla: Path, programa: Path, salida: Path) -> dict:
    contenido, resumen = generar_daily_check(plantilla, programa)
    salida.write_bytes(contenido)
    return resumen
